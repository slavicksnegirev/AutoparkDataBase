import pandas as pd
import openpyxl
import mysql.connector

from mysql.connector import Error



class DataBase():
    def __init__(self):
        super(DataBase, self).__init__()
        self.connection = self.create_connection("localhost", "root", "142536475869", "mydb")
        self.create_database(self.connection, "USE mydb")

    def create_connection(self, host_name, user_name, user_password, db_name):
        connection = None
        try:
            connection = mysql.connector.connect(
                host=host_name,
                user=user_name,
                passwd=user_password,
                database=db_name
            )
            print("Connection to MySQL DB successful")
        except Error as e:
            print(f"The error '{e}' occurred")

        return connection

    def create_database(self, connection, query):
        cursor = connection.cursor()
        try:
            cursor.execute(query)
            print("Database created successfully")
        except Error as e:
            print(f"The error '{e}' occurred")

    def execute_query(self, connection, query):
        cursor = self.connection.cursor(buffered=True)
        try:
            cursor.execute(query)
            self.connection.commit()
            print("Query executed successfully")
        except Error as e:
            print(f"The error '{e}' occurred")

    def execute_read_query(self, query):
        cursor = self.connection.cursor()
        result = None
        try:
            cursor.execute(query)
            result = cursor.fetchall()
            return result
        except Error as e:
            print(f"The error '{e}' occurred")


    def load_data_to_xlsx(self, data):
        work_book = openpyxl.Workbook()
        sheet = work_book.active

        for i in range(1, 100):
            for j in range(1, 100):
                sheet.cell(row=i, column=j).value = ''

        for row, item in enumerate(data):
            for j in range(len(item)):
                sheet.cell(row=row+1, column=j+1).value = str(item[j])


        work_book.save('output.xlsx')
        work_book.close()

    def insert_driver_SSI(self, surname, name, patronymic, date_of_birth, address, phone_number,
                          license_number):
        query = f"""INSERT INTO driver_SSI (`surname`, `name`, `patronymic`, `date_of_birth`, `address`, 
        `phone_number`, `license_number`) VALUES ("{surname}", "{name}", "{patronymic}", "{date_of_birth}", "{address}", "{phone_number}", 
        "{license_number}")"""
        # cursor = self.connection.cursor()
        # cursor.executemany(query, [(surname, name, patronymic, date_of_birth, address, phone_number, license_number)])
        # self.connection.commit()
        self.execute_query(self.connection, query)

    def edit_driver_SSI(self, surname, name, patronymic, date_of_birth, address, phone_number,
                        license_number, driver_ID):
        query = f"""UPDATE driver_SSI SET surname = "{surname}", name = "{name}", patronymic = "{patronymic}", date_of_birth =
         "{date_of_birth}", address = "{address}", phone_number = "{phone_number}", license_number = "{license_number}" WHERE driver_ID = {driver_ID}"""
        self.execute_query(self.connection, query)

    def delete_driver_SSI(self, driver_ID):
        query = f"""DELETE FROM driver_SSI WHERE driver_ID = {driver_ID}"""
        self.execute_query(self.connection, query)

    def get_driver_IDs(self):
        query = "SELECT driver_ID FROM driver_SSI"
        result = self.execute_read_query(query)

        return result

    def insert_automobile_SSI(self, license_plate, brand, color, mileage, year_of_release, engine_power, maximum_speed, fuel_consumption):
        query = f"""INSERT INTO automobile_SSI (`license_plate`, `brand`, `color`, `mileage`, `year_of_release`, 
        `engine_power`, `maximum_speed`, `fuel_consumption`) VALUES ("{license_plate}", "{brand}", "{color}", 
        "{mileage}", "{year_of_release}", "{engine_power}", "{maximum_speed}", "{fuel_consumption}" )"""
        self.execute_query(self.connection, query)

    def edit_automobile_SSI(self, license_plate, brand, color, mileage, year_of_release, engine_power, maximum_speed,
                            fuel_consumption, automobile_ID):
        query = f"""UPDATE automobile_SSI SET license_plate = "{license_plate}", brand = "{brand}", color = "{color}", mileage = "{mileage}", year_of_release = "{year_of_release}", engine_power = "{engine_power}",
        maximum_speed = "{maximum_speed}", fuel_consumption = "{fuel_consumption}" WHERE automobile_ID = {automobile_ID}"""
        self.execute_query(self.connection, query)

    def delete_automobile_SSI(self, automobile_ID):
        query = f"""DELETE FROM automobile_SSI WHERE automobile_ID = {automobile_ID}"""
        self.execute_query(self.connection, query)

    def get_automobile_IDs(self):
        query = "SELECT automobile_ID FROM automobile_SSI"
        result = self.execute_read_query(query)

        return result

    def insert_mechanic_SSI(self, surname, name, patronymic, date_of_birth, address, phone_number):
        query = f"""INSERT INTO mechanic_SSI (`surname`, `name`, `patronymic`, `date_of_birth`, `address`, `phone_number`) VALUES ("{surname}", "{name}", "{patronymic}", "{date_of_birth}", "{address}", "{phone_number}") """
        self.execute_query(self.connection, query)

    def edit_mechanic_SSI(self, surname, name, patronymic, date_of_birth, address, phone_number, mechanic_ID):
        query = f"""UPDATE mechanic_SSI SET surname = "{surname}", name = "{name}", patronymic = "{patronymic}", date_of_birth =
         "{date_of_birth}", address = "{address}", phone_number = "{phone_number}" WHERE mechanic_ID = {mechanic_ID}"""
        self.execute_query(self.connection, query)

    def delete_mechanic_SSI(self, mechanic_ID):
        query = f"""DELETE FROM mechanic_SSI WHERE mechanic_ID = {mechanic_ID}"""
        self.execute_query(self.connection, query)

    def get_mechanic_IDs(self):
        query = "SELECT mechanic_ID FROM mechanic_SSI"
        result = self.execute_read_query(query)

        return result

    def insert_client_SSI(self, surname, name, patronymic, phone_number):
        query = f"""INSERT INTO client_SSI (`surname`, `name`, `patronymic`, `phone_number`) VALUES ("{surname}", "{name}", "{patronymic}", "{phone_number}")"""
        self.execute_query(self.connection, query)

    def edit_client_SSI(self, surname, name, patronymic, phone_number, client_ID):
        query = f"""UPDATE client_SSI SET surname = "{surname}", name = "{name}", patronymic = "{patronymic}", 
        phone_number = "{phone_number}" WHERE client_ID = {client_ID}"""
        self.execute_query(self.connection, query)

    def delete_client_SSI(self, client_ID):
        query = f"""DELETE FROM client_SSI WHERE client_ID = {client_ID}"""
        self.execute_query(self.connection, query)

    def get_client_IDs(self):
        query = "SELECT client_ID FROM client_SSI"
        result = self.execute_read_query(query)

        return result

    def insert_order_SSI(self, datetime, description, departure_address, delivery_address, client_ID, trip_ID):
        query = f"""INSERT INTO order_SSI (`datetime`, `description`, `departure_address`, `delivery_address`, 
        `client_ID`, `trip_ID`) VALUES ("{datetime}", "{description}", "{departure_address}", 
        "{delivery_address}", "{client_ID}", "{trip_ID}")"""
        self.execute_query(self.connection, query)

    def edit_order_SSI(self, datetime, description, departure_address, delivery_address, client_ID, trip_ID, order_ID):
        self.execute_query(self.connection, "SET FOREIGN_KEY_CHECKS=0")
        query = f"""UPDATE order_SSI SET datetime = "{datetime}", description = "{description}", departure_address = "{departure_address}", delivery_address = "{delivery_address}",  client_ID = "{client_ID}", trip_ID = "{trip_ID}" WHERE order_ID = {order_ID}"""
        self.execute_query(self.connection, query)
        self.execute_query(self.connection, "SET FOREIGN_KEY_CHECKS=1")

    def delete_order_SSI(self, order_ID):
        query = f"""DELETE FROM order_SSI WHERE order_ID = {order_ID}"""
        self.execute_query(self.connection, query)

    def get_order_IDs(self):
        query = "SELECT order_ID FROM order_SSI"
        result = self.execute_read_query(query)

        return result

    def insert_trip_SSI(self, date, departure_time, arrival_time, mileage, driver_ID, automobile_ID):
        query = f"""INSERT INTO trip_SSI (`date`, `departure_time`, `arrival_time`, `mileage`, `driver_ID`, 
        `automobile_ID`) VALUES ("{date}", "{departure_time}", "{arrival_time}", "{mileage}", "{driver_ID}", "{automobile_ID}")"""
        self.execute_query(self.connection, query)

    def edit_trip_SSI(self, date, departure_time, arrival_time, mileage, driver_ID, automobile_ID, trip_ID):
        query = f"""UPDATE trip_SSI SET date = "{date}", departure_time = "{departure_time}", arrival_time = 
        "{arrival_time}", mileage = "{mileage}", driver_ID = "{driver_ID}", automobile_ID = 
        "{automobile_ID}" WHERE trip_ID = {trip_ID}"""
        self.execute_query(self.connection, query)

    def delete_trip_SSI(self, trip_ID):
        query = f"""DELETE FROM trip_SSI WHERE trip_ID = {trip_ID}"""
        self.execute_query(self.connection, query)

    def get_trip_IDs(self):
        query = "SELECT trip_ID FROM trip_SSI"
        result = self.execute_read_query(query)

        return result

    def insert_service_SSI(self, datetime, title, description, automobile_ID, mechanic_ID):
        query = f"""INSERT INTO service_SSI (`datetime`, `title`, `description`, `automobile_ID`, `mechanic_ID`) 
        VALUES ("{datetime}", "{title}", "{description}", "{automobile_ID}", "{mechanic_ID}")"""
        self.execute_query(self.connection, query)

    def edit_service_SSI(self, datetime, title, description, automobile_ID, mechanic_ID, service_ID):
        query = f"""UPDATE service_SSI SET datetime = "{datetime}", title = "{title}", description = 
        "{description}", automobile_ID = "{automobile_ID}", mechanic_ID = "{mechanic_ID}" WHERE service_ID = {service_ID}"""
        self.execute_query(self.connection, query)

    def delete_service_SSI(self, service_ID):
        query = f"""DELETE FROM service_SSI WHERE service_ID = {service_ID}"""
        self.execute_query(self.connection, query)

    def get_service_IDs(self):
        query = "SELECT service_ID FROM service_SSI"
        result = self.execute_read_query(query)

        return result

    def insert_authorization_SSI(self, login, password):
        query = f"""INSERT INTO authorization_SSI (`login`, `password`) VALUES ("{login}", "{password}")"""
        self.execute_query(self.connection, query)

    def edit_authorization_SSI(self, login, password, authorization_ID):
        query = f"""UPDATE authorization_SSI SET login = "{login}", password = "{password}" WHERE authorization_ID =
         {authorization_ID}"""
        self.execute_query(self.connection, query)

    def delete_authorization_SSI(self, authorization_ID):
        query = f"""DELETE FROM authorization_SSI WHERE authorization_ID = {authorization_ID}"""
        self.execute_query(self.connection, query)

    def get_authorization_IDs(self):
        query = "SELECT authorization_ID FROM authorization_SSI"
        result = self.execute_read_query(query)

        return result



    def get_all_workers_lastname(self):
        query = """SELECT concat(surname, SPACE(1), name, SPACE(1), patronymic) as fullname
        FROM driver_SSI
        UNION SELECT concat(surname, SPACE(1), name, SPACE(1), patronymic) as fullname FROM mechanic_SSI;
        """

        result = self.execute_read_query(query)

        return result

    def check_authorization(self, login, password):
        query = f"""SELECT EXISTS(SELECT * FROM authorization_SSI WHERE login = '{login}' and password = '{password}')"""
        return self.execute_read_query(query)


    def report_1(self):
        query = f"""
        select date,
        concat(dS.surname, space(1), dS.name, space(1), dS.patronymic) as driver,
        concat(aSS.brand, space(1), aSS.color, space(1), aSS.year_of_release) as car
        from order_SSI oS
        left join trip_SSI tS on ts.trip_ID = oS.trip_ID
        left join automobile_SSI aSS on tS.automobile_ID = aSS.automobile_ID
        left join driver_SSI dS on tS.driver_ID = dS.driver_ID
        """
        return self.execute_read_query(query)

    def report_2(self):
        query = f"""
        select date,
        concat(driver_SSI.surname, space(1), driver_SSI.name, space(1), driver_SSI.patronymic) as driver,
        concat(automobile_SSI.brand, space(1), automobile_SSI.color, space(1), automobile_SSI.year_of_release) as car
        from trip_SSI
        join driver_SSI on trip_SSI.driver_ID = driver_SSI.driver_ID
        join automobile_SSI on trip_SSI.automobile_ID = automobile_SSI.automobile_ID;
        """
        return self.execute_read_query(query)

    def report_3(self):
        query = f"""
        select datetime,
        concat(mechanic_SSI.surname, space(1), mechanic_SSI.name, space(1), mechanic_SSI.patronymic) as mechanic,
        concat(automobile_SSI.brand, space(1), automobile_SSI.color, space(1), automobile_SSI.year_of_release) as car
        from service_SSI
        join mechanic_SSI on service_SSI.mechanic_ID = mechanic_SSI.mechanic_ID
        join automobile_SSI on service_SSI.automobile_ID = automobile_SSI.automobile_ID;
        """
        return self.execute_read_query(query)